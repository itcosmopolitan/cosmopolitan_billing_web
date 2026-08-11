import json
import uuid
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import String, and_, asc, cast, delete, func, or_, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from src.batch_dates import validate_batch_dates
from src.database import get_db
from src.item_branch import effective_cost_price, effective_reorder_level, effective_selling_price
from src.models import Branch, Category, Item, ItemApprovalStatus, ItemBatch, ItemBranchConfig, ItemStock, User, AuditLog
from src.models import (
    AdjustmentRequest,
    AdjustmentStatus,
    Branch,
    Category,
    GoodsReceiptNote,
    GRNLineItem,
    Item,
    ItemBatch,
    ItemBranchConfig,
    ItemStock,
    PurchaseBill,
    PurchaseLineItem,
    PurchaseOrder,
    PurchaseOrderLineItem,
    Quotation,
    QuotationLineItem,
    ReturnLineItem,
    SaleInvoice,
    SaleLineItem,
    SalesOrder,
    SalesOrderLineItem,
    SalesReturn,
    SalesReturnLineItem,
    StockAdjustment,
    StockMovement,
    StockReservation,
    StockTransfer,
    TransferLineItem,
    TransferStatus,
)
from src.pagination import (
    normalize_limit,
    normalize_skip,
    paged,
    pagination_from_page,
    resolve_sort,
)
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
import re
from src.routes._atomic import (
    add_batch_atomic,
    adjust_stock_atomic,
    consume_batches_atomic,
    set_batch_quantity_atomic,
    set_stock_atomic,
)
from src.permissions import ITEM_CATALOG_READ
from src.routes._approval import can_direct_commit
from src.routes._approval import user_can
from src.security import (
    current_user,
    enforce_branch_access_optional,
    enforce_branch_access,
    get_allowed_branch_ids,
    require_perm,
)
from src.services.audit_service import build_audit_entry

router = APIRouter()


async def _write_post_commit_audit(
    db: AsyncSession,
    *,
    action: str,
    reference_id: str,
    detail: str,
    user: Optional[User],
    request: Request = None,
    branch_id: Optional[str],
    metadata: Optional[dict] = None,
) -> None:
    role = "unknown"
    if user is not None:
        role = user.role.value if hasattr(user.role, "value") else str(user.role)
    payload = build_audit_entry(
        action=action,
        module="Inventory",
        reference_id=reference_id,
        detail=detail,
        user_id=user.id if user is not None else "system",
        user_name=user.name if user is not None else "System",
        user_role=role,
        ip_address=getattr(request.state, "ip_address", None),
        device_info=getattr(request.state, "device_info", None),
        branch_id=branch_id,
        metadata=metadata,
    )
    db.add(AuditLog(id=str(uuid.uuid4()), **payload))
    await db.commit()


def _log_item_history(
    db: AsyncSession,
    *,
    user: Optional[User],
    item: Item,
    event_type: str,
    action: str,
    detail: str,
    metadata: Optional[dict] = None,
    risk: str = "low",
) -> None:
    role = "unknown"
    if user is not None:
        role = user.role.value if hasattr(user.role, "value") else str(user.role)

    db.add(AuditLog(
        id=str(uuid.uuid4()),
        record_type="item",
        record_id=item.id,
        event_type=event_type,
        event_metadata=json.dumps(metadata or {}, default=str),
        action=action,
        user_id=getattr(user, "id", None),
        user_name=getattr(user, "name", None),
        user_role=role,
        module="inventory",
        reference_id=item.sku or item.name,
        ref=item.sku or item.name,
        detail=detail,
        risk=risk,
        ip_address=None,
    ))


def _resolve_item_status_value(item: Item) -> str:
    for attr in ("status", "approval_status"):
        value = getattr(item, attr, None)
        if value is None:
            continue
        if hasattr(value, "value"):
            value = value.value
        else:
            value = str(value)
        if value in {ItemApprovalStatus.pending.value, ItemApprovalStatus.pending_approval.value}:
            return ItemApprovalStatus.pending.value
        return value
    return "approved"


async def _get_config_map(
    db: AsyncSession,
    item_ids: list[str],
    branch_id: Optional[str],
) -> dict[str, ItemBranchConfig]:
    if not item_ids or not branch_id:
        return {}
    res = await db.execute(
        select(ItemBranchConfig).where(
            ItemBranchConfig.item_id.in_(item_ids),
            ItemBranchConfig.branch_id == branch_id,
        )
    )
    return {row.item_id: row for row in res.scalars().all()}


async def _available_branch_counts(
    db: AsyncSession,
    item_ids: list[str],
) -> dict[str, int]:
    if not item_ids:
        return {}
    res = await db.execute(
        select(ItemBranchConfig.item_id, func.count(ItemBranchConfig.id))
        .where(
            ItemBranchConfig.item_id.in_(item_ids),
            ItemBranchConfig.is_available == True,  # noqa: E712
        )
        .group_by(ItemBranchConfig.item_id)
    )
    return {item_id: int(cnt or 0) for item_id, cnt in res.all()}


async def _upsert_branch_config(
    db: AsyncSession,
    *,
    item_id: str,
    branch_id: str,
    is_available: bool = True,
    cost_price: Optional[float] = None,
    selling_price: Optional[float] = None,
    reorder_level: Optional[int] = None,
) -> ItemBranchConfig:
    res = await db.execute(
        select(ItemBranchConfig).where(
            ItemBranchConfig.item_id == item_id,
            ItemBranchConfig.branch_id == branch_id,
        )
    )
    cfg = res.scalar_one_or_none()
    if cfg:
        cfg.is_available = is_available
        cfg.cost_price = cost_price
        cfg.selling_price = selling_price
        cfg.reorder_level = reorder_level
        return cfg
    cfg = ItemBranchConfig(
        id=str(uuid.uuid4()),
        item_id=item_id,
        branch_id=branch_id,
        is_available=is_available,
        cost_price=cost_price,
        selling_price=selling_price,
        reorder_level=reorder_level,
    )
    db.add(cfg)
    return cfg


async def _ensure_stock_row(
    db: AsyncSession,
    *,
    item_id: str,
    branch_id: str,
) -> None:
    await db.execute(
        text(
            "INSERT INTO item_stock (id, item_id, branch_id, quantity) "
            "VALUES (:id, :item_id, :branch_id, 0) "
            "ON CONFLICT(item_id, branch_id) DO NOTHING"
        ),
        {
            "id": str(uuid.uuid4()),
            "item_id": item_id,
            "branch_id": branch_id,
        },
    )


async def _stock_qty(db: AsyncSession, *, item_id: str, branch_id: str) -> int:
    res = await db.execute(
        select(ItemStock.quantity).where(
            ItemStock.item_id == item_id,
            ItemStock.branch_id == branch_id,
        )
    )
    return int(res.scalar() or 0)


async def _seed_opening_stock(
    db: AsyncSession,
    *,
    item: Item,
    branch_id: str,
    qty: int,
    cost_price: float,
    batch_number: Optional[str] = None,
    mfg_date: Optional[str] = None,
    expiry_date: Optional[str] = None,
) -> None:
    """Add opening qty at a branch (create or edit). Tracked → new batch lot; untracked → bump stock."""
    if qty <= 0:
        return
    await _ensure_stock_row(db, item_id=item.id, branch_id=branch_id)
    if item.batch_tracking:
        _raise_batch_date_errors(validate_batch_dates(
            mfg_date=mfg_date,
            expiry_date=expiry_date,
            require_expiry=bool(item.expiry_tracking),
        ))
        await add_batch_atomic(
            db,
            item_id=item.id,
            branch_id=branch_id,
            qty=int(qty),
            batch_number=batch_number,
            mfg_date=mfg_date,
            expiry_date=expiry_date,
            cost_price=cost_price,
            source_type="opening",
            source_ref=item.id,
            received_date=mfg_date or None,
        )
    else:
        await adjust_stock_atomic(
            db,
            item_id=item.id,
            branch_id=branch_id,
            delta=int(qty),
        )


def _raise_batch_date_errors(errors: list[str]) -> None:
    if errors:
        raise HTTPException(400, "; ".join(errors))


async def _apply_batch_tracking_toggle(
    db: AsyncSession,
    item: Item,
    *,
    previously_tracked: bool,
    enable: bool,
) -> Optional[dict]:
    """Reconcile ``item_batches`` when ``batch_tracking`` flips.

    OFF → delete all batch rows; ``item_stock`` stays the source of truth
    (untracked adjustments / POS sales already kept aggregate correct).

    ON  → delete any stale batch rows, then seed one opening batch per branch
    from current ``item_stock`` *without* bumping aggregate (avoids double
    counting vs ``add_batch_atomic``).
    """
    if previously_tracked == enable:
        return None

    if not enable:
        res = await db.execute(
            delete(ItemBatch).where(ItemBatch.item_id == item.id)
        )
        item.expiry_tracking = False
        return {
            "action": "disabled",
            "batches_deleted": int(res.rowcount or 0),
        }

    await db.execute(delete(ItemBatch).where(ItemBatch.item_id == item.id))
    stock_rows = (
        await db.execute(select(ItemStock).where(ItemStock.item_id == item.id))
    ).scalars().all()
    today = datetime.utcnow().strftime("%Y-%m-%d")
    seeded = 0
    for st in stock_rows:
        qty = int(st.quantity or 0)
        if qty <= 0:
            continue
        db.add(ItemBatch(
            id=str(uuid.uuid4()),
            item_id=item.id,
            branch_id=st.branch_id,
            batch_number=f"OPEN-{uuid.uuid4().hex[:8].upper()}",
            quantity=qty,
            initial_qty=qty,
            cost_price=float(item.cost_price or 0),
            source_type="opening",
            source_ref=item.id,
            received_date=today,
            notes="Opening batch created when batch tracking was enabled",
            active=True,
        ))
        seeded += 1
    if seeded:
        await db.flush()
    return {"action": "enabled", "batches_seeded": seeded}

async def _item_inventory_blockers(
    db: AsyncSession,
    item_id: str,
    branch_id: Optional[str] = None,
) -> list[str]:
    """Check blockers for item deletion/deactivation.
    
    If branch_id is provided, only check that branch's stock and transactions.
    If branch_id is None, check globally (for Item Master operations).
    """
    reasons: list[str] = []

    if branch_id:
        # Branch-specific checks (Items & Stock page context)
        stock_res = await db.execute(
            select(ItemStock.quantity).where(
                ItemStock.item_id == item_id,
                ItemStock.branch_id == branch_id,
            )
        )
        qty = int(stock_res.scalar() or 0)
        if qty > 0:
            reasons.append(f"branch has {qty} units in stock")

        batch_res = await db.execute(
            select(func.sum(ItemBatch.quantity)).where(
                ItemBatch.item_id == item_id,
                ItemBatch.branch_id == branch_id,
            )
        )
        if int(batch_res.scalar() or 0) > 0:
            reasons.append("branch has non-zero batch quantity")

        pending_adj = await db.execute(
            select(AdjustmentRequest.id).where(
                AdjustmentRequest.item_id == item_id,
                AdjustmentRequest.branch_id == branch_id,
                AdjustmentRequest.status == AdjustmentStatus.pending,
            ).limit(1)
        )
        if pending_adj.scalar_one_or_none():
            reasons.append("branch has pending stock adjustment requests")

        transfer_exists = await db.execute(
            select(StockTransfer.id)
            .join(TransferLineItem, TransferLineItem.transfer_id == StockTransfer.id)
            .where(
                TransferLineItem.item_id == item_id,
                or_(
                    StockTransfer.from_branch_id == branch_id,
                    StockTransfer.to_branch_id == branch_id,
                ),
                StockTransfer.status.in_(
                    [TransferStatus.pending, TransferStatus.approved, TransferStatus.transit]
                ),
            )
            .limit(1)
        )
        if transfer_exists.scalar_one_or_none():
            reasons.append("branch has pending or active stock transfer requests")
    else:
        # Global checks (Item Master context)
        stock_res = await db.execute(
            select(func.sum(ItemStock.quantity)).where(ItemStock.item_id == item_id)
        )
        if int(stock_res.scalar() or 0) > 0:
            reasons.append("item has non-zero stock")

        batch_res = await db.execute(
            select(func.sum(ItemBatch.quantity)).where(ItemBatch.item_id == item_id)
        )
        if int(batch_res.scalar() or 0) > 0:
            reasons.append("item has non-zero batch quantity")

        pending_adj = await db.execute(
            select(AdjustmentRequest.id).where(
                AdjustmentRequest.item_id == item_id,
                AdjustmentRequest.status == AdjustmentStatus.pending,
            ).limit(1)
        )
        if pending_adj.scalar_one_or_none():
            reasons.append("item has pending stock adjustment requests")

        transfer_exists = await db.execute(
            select(StockTransfer.id)
            .join(TransferLineItem, TransferLineItem.transfer_id == StockTransfer.id)
            .where(
                TransferLineItem.item_id == item_id,
                StockTransfer.status.in_(
                    [TransferStatus.pending, TransferStatus.approved, TransferStatus.transit]
                ),
            )
            .limit(1)
        )
        if transfer_exists.scalar_one_or_none():
            reasons.append("item has pending or active stock transfer requests")

    return reasons

async def _item_has_historical_references(
    db: AsyncSession,
    item_id: str,
    branch_id: Optional[str] = None,
) -> bool:
    """Check if item has historical transaction references.

    If branch_id is provided, only check that branch's transactions.
    If branch_id is None, check globally.
    """
    if branch_id:
        branch_checks = [
            (SaleLineItem, SaleInvoice, SaleLineItem.invoice_id, SaleInvoice.branch_id),
            (PurchaseLineItem, PurchaseBill, PurchaseLineItem.bill_id, PurchaseBill.branch_id),
            (QuotationLineItem, Quotation, QuotationLineItem.quotation_id, Quotation.branch_id),
            (SalesReturnLineItem, SalesReturn, SalesReturnLineItem.return_id, SalesReturn.branch_id),
        ]

        for line_model, parent_model, line_fk, parent_branch in branch_checks:
            result = await db.execute(
                select(line_model.id)
                .join(parent_model, line_fk == parent_model.id)
                .where(
                    line_model.item_id == item_id,
                    parent_branch == branch_id,
                ).limit(1)
            )
            if result.scalar_one_or_none():
                return True

        for model in (AdjustmentRequest, StockAdjustment):
            result = await db.execute(
                select(model.id).where(
                    model.item_id == item_id,
                    model.branch_id == branch_id,
                ).limit(1)
            )
            if result.scalar_one_or_none():
                return True

        return False

    for model in (
        SaleLineItem,
        SalesOrderLineItem,
        PurchaseLineItem,
        PurchaseOrderLineItem,
        QuotationLineItem,
        GRNLineItem,
        ReturnLineItem,
        SalesReturnLineItem,
        AdjustmentRequest,
        StockAdjustment,
        StockMovement,
        StockReservation,
        TransferLineItem,
    ):
        result = await db.execute(select(model.id).where(model.item_id == item_id).limit(1))
        if result.scalar_one_or_none():
            return True

    return False


async def _get_branch_batch_info(
    db: AsyncSession,
    item_id: str,
    branch_id: str,
) -> dict:
    """Get info about batches and stock at a branch for an item.
    
    Returns dict with:
    - batch_count: number of batches
    - batch_qty: total quantity in batches
    - stock_qty: quantity in item_stock
    """
    # Count batches
    batch_res = await db.execute(
        select(func.count(ItemBatch.id), func.sum(ItemBatch.quantity)).where(
            ItemBatch.item_id == item_id,
            ItemBatch.branch_id == branch_id,
        )
    )
    batch_count, batch_qty = batch_res.one()
    
    # Get stock quantity
    stock_res = await db.execute(
        select(ItemStock.quantity).where(
            ItemStock.item_id == item_id,
            ItemStock.branch_id == branch_id,
        )
    )
    stock_qty = stock_res.scalar() or 0
    
    return {
        "batch_count": int(batch_count or 0),
        "batch_qty": int(batch_qty or 0),
        "stock_qty": int(stock_qty or 0),
    }
class ItemCreate(BaseModel):
    name: str
    sku: Optional[str] = None
    barcode: Optional[str] = None
    country_of_origin: Optional[str] = None
    category_id: Optional[str] = None
    brand: Optional[str] = None
    is_packaging: bool = False
    packaging_quantity: Optional[float] = None
    unit: str = "Pcs"
    cost_price: float
    selling_price: float
    wholesale_discount_pct: float = 0
    staff_discount_pct: float = 0
    tax_rate: float = 18
    hsn_code: Optional[str] = None
    reorder_level: int = 10
    emoji: str = "📦"
    batch_tracking: bool = False
    expiry_tracking: bool = False
    active: Optional[bool] = None
    branch_id: Optional[str] = None
    opening_stock: int = 0
    branch_id: str = "br-001"
    # Optional opening-batch metadata. Honored only when batch_tracking is
    # true; ignored on legacy untracked items so the existing UI keeps working.
    opening_batch_number: Optional[str] = None
    opening_mfg_date:     Optional[str] = None
    opening_expiry_date:  Optional[str] = None
    # Per-branch listing / pricing. When omitted, only ``branch_id`` is listed.
    branch_configs: Optional[list["BranchConfigIn"]] = None


class BranchConfigIn(BaseModel):
    branch_id: str
    is_available: bool = True
    cost_price: Optional[float] = None
    selling_price: Optional[float] = None
    reorder_level: Optional[int] = None
    opening_stock: int = 0
    opening_batch_number: Optional[str] = None
    opening_mfg_date: Optional[str] = None
    opening_expiry_date: Optional[str] = None


class BranchConfigBulk(BaseModel):
    branches: list[BranchConfigIn]

class StockAdjustRequest(BaseModel):
    item_id: str
    branch_id: str
    new_qty: int
    reason: str
    notes: Optional[str] = None
    # When provided, the adjustment targets a single batch (per-batch absolute
    # set). When omitted on a batch-tracked item, the aggregate is changed and
    # — for a *decrease* — drawn from oldest batches first (FIFO/FEFO).
    batch_id: Optional[str] = None


class BatchCreate(BaseModel):
    branch_id: str
    qty: int
    batch_number: Optional[str] = None
    mfg_date:    Optional[str] = None
    expiry_date: Optional[str] = None
    cost_price:  float = 0
    vendor_id:   Optional[str] = None
    received_date: Optional[str] = None
    notes:       Optional[str] = None


class BatchPatch(BaseModel):
    """Partial edit for batch metadata (cannot change quantity here — use
    stock-adjust instead so the aggregate stays in sync)."""
    batch_number:  Optional[str] = None
    mfg_date:      Optional[str] = None
    expiry_date:   Optional[str] = None
    notes:         Optional[str] = None
    active:        Optional[bool] = None


class ItemPatch(BaseModel):
    """Partial item update body. Excludes immutable / derived fields like
    `id`, `sku` (managed by create), `active` (use a separate deactivation
    endpoint when added). `data: dict` + setattr would have allowed any
    column to be flipped from a PATCH."""
    name: Optional[str] = None
    barcode: Optional[str] = None
    country_of_origin: Optional[str] = None
    category_id: Optional[str] = None
    is_packaging: Optional[bool] = None
    packaging_quantity: Optional[float] = None
    brand: Optional[str] = None
    unit: Optional[str] = None
    cost_price: Optional[float] = None
    selling_price: Optional[float] = None
    wholesale_discount_pct: Optional[float] = None
    staff_discount_pct: Optional[float] = None
    tax_rate: Optional[float] = None
    hsn_code: Optional[str] = None
    reorder_level: Optional[int] = None
    emoji: Optional[str] = None
    batch_tracking: Optional[bool] = None
    expiry_tracking: Optional[bool] = None
    # Allow branch-specific deactivate/activate requests and toggling active
    # via PATCH: include `active` and `branch_id` so the route's deactivation
    # blocker logic receives the intended values.
    active: Optional[bool] = None
    branch_id: Optional[str] = None



class InventoryCategoryCreate(BaseModel):
    name: str
    icon: Optional[str] = "📦"


@router.get("/categories", dependencies=[Depends(require_perm("items.view", "item_master.view", "item_master.create", "item_master.edit"))])
async def list_categories(db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(Category).order_by(asc(Category.name)))).scalars().all()
    return [
        {"id": row.id, "name": row.name, "icon": row.icon or "📦"}
        for row in rows
    ]


@router.post("/categories", dependencies=[Depends(require_perm("item_master.create", "item_master.edit"))], status_code=201)
async def create_category(data: InventoryCategoryCreate, db: AsyncSession = Depends(get_db)):
    name = (data.name or "").strip()
    if not name:
        raise HTTPException(400, "Category name is required")

    existing = (
        await db.execute(select(Category).where(func.lower(Category.name) == name.lower()))
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(400, "Category already exists")

    category = Category(
        id=f"cat-{uuid.uuid4().hex[:8]}",
        name=name,
        icon=data.icon or "📦",
    )
    db.add(category)
    await db.commit()
    return {"id": category.id, "name": category.name, "icon": category.icon, "message": "Category created"}

@router.get("/", dependencies=[Depends(require_perm(*ITEM_CATALOG_READ))])
async def list_items(
    search: Optional[str] = None,
    category_id: Optional[str] = None,
    branch_id: Optional[str] = Depends(enforce_branch_access_optional),
    status: Optional[str] = None,
    include_inactive: bool = False,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "asc",
    page_no: Optional[int] = Query(None, ge=1),
    per_page: Optional[int] = Query(None, ge=1, le=500),
    skip: Optional[int] = Query(None, ge=0),
    limit: Optional[int] = Query(None, ge=1, le=500),
    include_total: bool = True,
    pos_mode: bool = False,
    master_mode: bool = False,
    listed_only: Optional[bool] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    if page_no is not None or per_page is not None:
        pn, pp, sk, lim = pagination_from_page(page_no, per_page)
    else:
        sk = normalize_skip(skip)
        lim = normalize_limit(limit)
        pn = max(1, (sk // lim) + 1)
        pp = lim

    filter_listed = listed_only if listed_only is not None else (pos_mode or not master_mode)

    if branch_id is None and not getattr(user, "all_branches", False) and not master_mode:
        branch_ids = await get_allowed_branch_ids(user, db)
        if not branch_ids:
            return paged([], 0, sk, lim)
        branch_id = branch_ids[0]

    q = select(Item).options(selectinload(Item.category))
    cq = select(func.count(Item.id))

    if master_mode and status == "pending":
        pending_filter = or_(
            cast(Item.approval_status, String) == ItemApprovalStatus.pending.value,
            cast(Item.approval_status, String) == ItemApprovalStatus.pending_approval.value,
        )
        q = q.where(pending_filter)
        cq = cq.where(pending_filter)
    elif master_mode and status == "rejected":
        q = q.where(cast(Item.approval_status, String) == ItemApprovalStatus.rejected.value)
        cq = cq.where(cast(Item.approval_status, String) == ItemApprovalStatus.rejected.value)
    elif master_mode and status == "inactive":
        q = q.where(
            Item.active.is_(False),
            cast(Item.approval_status, String) != ItemApprovalStatus.rejected.value,
            cast(Item.approval_status, String) != ItemApprovalStatus.pending_approval.value,
            cast(Item.approval_status, String) != ItemApprovalStatus.pending.value,
        )
        cq = cq.where(
            Item.active.is_(False),
            cast(Item.approval_status, String) != ItemApprovalStatus.rejected.value,
            cast(Item.approval_status, String) != ItemApprovalStatus.pending_approval.value,
            cast(Item.approval_status, String) != ItemApprovalStatus.pending.value,
        )
    elif master_mode and status in {None, "approved", "all"}:
        q = q.where(cast(Item.approval_status, String) == ItemApprovalStatus.approved.value)
        cq = cq.where(cast(Item.approval_status, String) == ItemApprovalStatus.approved.value)
    elif not master_mode or pos_mode:
        q = q.where(cast(Item.approval_status, String) == ItemApprovalStatus.approved.value)
        cq = cq.where(cast(Item.approval_status, String) == ItemApprovalStatus.approved.value)
    # Note: do not pre-filter by `active` here — apply the active/inactive
    # filter later based on the `include_inactive` flag so callers that set
    # `include_inactive=true` actually receive inactive rows. The previous
    # unconditional filter caused the Inactive tab to never show deactivated
    # items.

    if not include_inactive and status not in {"pending", "rejected"}:
        q = q.where(Item.active == True)  # noqa: E712
        cq = cq.where(Item.active == True)  # noqa: E712

    if branch_id:
        listed_join = and_(
            ItemBranchConfig.item_id == Item.id,
            ItemBranchConfig.branch_id == branch_id,
            ItemBranchConfig.is_available == True,  # noqa: E712
        )
        if filter_listed:
            listed_join = and_(
                listed_join,
                ItemBranchConfig.is_available == True,  # noqa: E712
            )
        q = q.join(ItemBranchConfig, listed_join)
        cq = cq.join(ItemBranchConfig, listed_join)

    if search:
        term = f"%{search}%"
        q = q.where(
            (Item.name.ilike(term)) | (Item.sku.ilike(term)) | (Item.barcode.ilike(term))
        )
        cq = cq.where(
            (Item.name.ilike(term)) | (Item.sku.ilike(term)) | (Item.barcode.ilike(term))
        )
    if category_id:
        q = q.where(Item.category_id == category_id)
        cq = cq.where(Item.category_id == category_id)
    total = 0
    if include_total and not pos_mode:
        total = int((await db.execute(cq)).scalar() or 0)
    # NB: `available_stock` is a per-branch lookup done in Python below, not a
    # column on Item — we can't sort by it at the SQL level without a join. If
    # that becomes important, restructure the query to join ItemStock and add
    # `available_stock` to the allow-list as a query expression.
    sort_expr = resolve_sort(
        sort_by,
        sort_order,
        {
            "name": Item.name,
            "sku": Item.sku,
            "barcode": Item.barcode,
            "category_id": Item.category_id,
            "cost_price": Item.cost_price,
            "selling_price": Item.selling_price,
            "tax_rate": Item.tax_rate,
            "reorder_level": Item.reorder_level,
            "created_at": Item.created_at,
            "updated_at": Item.updated_at,
        },
        default_key="name",
        default_order="asc",
    )
    fetch_limit = lim + 1 if (pos_mode and not include_total) else lim
    result = await db.execute(q.order_by(sort_expr).offset(sk).limit(fetch_limit))
    items = result.scalars().all()
    has_more = False
    if pos_mode and not include_total:
        has_more = len(items) > lim
        items = items[:lim]
    ids = [it.id for it in items]
    stock_by_item = {}
    if ids and branch_id:
        sr = await db.execute(
            select(ItemStock).where(and_(ItemStock.item_id.in_(ids), ItemStock.branch_id == branch_id))
        )
        for row in sr.scalars().all():
            stock_by_item[row.item_id] = row.quantity
    config_by_item = await _get_config_map(db, ids, branch_id)
    avail_branch_counts = await _available_branch_counts(db, ids) if master_mode else {}
    # Per-item batch summary (count + nearest expiry) so the Items table can
    # show a "🧴 N batches • exp dd MMM" hint for tracked items without an
    # extra round-trip per row.
    batch_counts: dict[str, int] = {}
    nearest_expiry: dict[str, str] = {}
    if ids:
        br = await db.execute(
            select(
                ItemBatch.item_id,
                func.count(ItemBatch.id),
                func.min(ItemBatch.expiry_date),
            )
            .where(
                ItemBatch.item_id.in_(ids),
                ItemBatch.branch_id == branch_id,
                ItemBatch.quantity > 0,
                ItemBatch.active == True,  # noqa: E712
            )
            .group_by(ItemBatch.item_id)
        )
        for item_id, cnt, exp in br.all():
            batch_counts[item_id] = int(cnt or 0)
            if exp:
                nearest_expiry[item_id] = exp
    out = []
    for item in items:
        cfg = config_by_item.get(item.id)
        eff_price = effective_selling_price(item.selling_price, cfg.selling_price if cfg else None)
        eff_cost = effective_cost_price(item.cost_price, cfg.cost_price if cfg else None)
        eff_reorder = effective_reorder_level(item.reorder_level, cfg.reorder_level if cfg else None)
        resolved_status = _resolve_item_status_value(item)
        row = {
            "id": item.id,
            "name": item.name,
            "sku": item.sku,
            "barcode": item.barcode,
            "categoryId": item.category_id,
            "categoryName": item.category.name if item.category else "Uncategorized",
            "brand": item.brand,
            "unit": item.unit,
            "default_cost_price": item.cost_price,
            "cost_price": item.cost_price if master_mode else eff_cost,
            "default_selling_price": item.selling_price,
            "selling_price": eff_price,
            "wholesale_discount_pct": float(getattr(item, "wholesale_discount_pct", 0) or 0),
            "staff_discount_pct": float(getattr(item, "staff_discount_pct", 0) or 0),
            "tax_rate": item.tax_rate,
            "hsn_code": item.hsn_code,
            "reorder_level": eff_reorder,
            "default_reorder_level": item.reorder_level,
            "emoji": item.emoji,
            "batch_tracking": item.batch_tracking,
            "expiry_tracking": item.expiry_tracking,
            "available_stock": stock_by_item.get(item.id, 0),
            "batches_count":   batch_counts.get(item.id, 0),
            "nearest_expiry":  nearest_expiry.get(item.id),
            "is_available": bool(cfg.is_available) if cfg else False,
            "branch_cost_override": cfg.cost_price if cfg else None,
            "branch_price_override": cfg.selling_price if cfg else None,
            "active": bool(item.active),
            "status": resolved_status,
            "approval_status": resolved_status,
        }
        if master_mode:
            row["available_branch_count"] = avail_branch_counts.get(item.id, 0)
            row["active"] = bool(item.active)
            row["created_by"] = item.created_by
            row["created_at"] = item.created_at.isoformat() if item.created_at else None
        out.append(row)
    if pos_mode and not include_total:
        return paged(
            out,
            total,
            sk,
            lim,
            page_context={
                "page_no": pn,
                "per_page": pp,
                "has_more_page": has_more,
            },
        )
    return paged(out, total, sk, lim)

async def _validate_category_id(category_id: Optional[str], db: AsyncSession) -> None:
    if not category_id:
        return
    result = await db.execute(select(Category).where(Category.id == category_id))
    if result.scalar_one_or_none() is None:
        raise HTTPException(400, detail=f"Invalid category_id: {category_id}")
    
@router.post("/", dependencies=[Depends(require_perm("item_master.create"))])
async def create_item(
    data: ItemCreate,
    request: Request = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    await _validate_category_id(data.category_id, db)
    if data.is_packaging and (data.packaging_quantity is None or data.packaging_quantity <= 0):
        raise HTTPException(400, "Packaging quantity must be greater than zero when packaging is enabled")
    direct = await can_direct_commit(user, db, "item_master.approve")
    initial_status = ItemApprovalStatus.approved if direct else ItemApprovalStatus.pending
    item = Item(
        id=str(uuid.uuid4()),
        name=data.name,
        sku=data.sku or f"SKU-{uuid.uuid4().hex[:6].upper()}",
        barcode=data.barcode,
        country_of_origin=data.country_of_origin,
        category_id=data.category_id or None,
        brand=data.brand,
        unit=data.unit,
        cost_price=data.cost_price,
        selling_price=data.selling_price,
        wholesale_discount_pct=float(data.wholesale_discount_pct or 0),
        staff_discount_pct=float(data.staff_discount_pct or 0),
        tax_rate=data.tax_rate,
        hsn_code=data.hsn_code,
        reorder_level=data.reorder_level,
        is_packaging=data.is_packaging,
        packaging_quantity=data.packaging_quantity if data.is_packaging else None,
        emoji=data.emoji,
        batch_tracking=data.batch_tracking,
        expiry_tracking=data.expiry_tracking,
        active=bool(direct),
        approval_status=initial_status,
        status=initial_status,
        created_by=user.name,
        approved_by=None,
        approved_at=None,
        rejected_by=None,
        rejected_at=None,
        rejection_reason=None,
    )
    db.add(item)
    await db.flush()

    configs = data.branch_configs

    if configs:
        seeded_any = False
        for bc in configs:
            branch_cost = effective_cost_price(data.cost_price, bc.cost_price)
            # Enforce that the operator may access the branch they're configuring
            await enforce_branch_access(bc.branch_id, user=user, db=db)
            await _upsert_branch_config(
                db,
                item_id=item.id,
                branch_id=bc.branch_id,
                is_available=bc.is_available,
                cost_price=bc.cost_price,
                selling_price=bc.selling_price,
                reorder_level=bc.reorder_level,
            )
            if bc.is_available:
                await _ensure_stock_row(db, item_id=item.id, branch_id=bc.branch_id)
            if bc.is_available and int(bc.opening_stock or 0) > 0:
                await _seed_opening_stock(
                    db,
                    item=item,
                    branch_id=bc.branch_id,
                    qty=int(bc.opening_stock),
                    cost_price=branch_cost,
                    batch_number=bc.opening_batch_number,
                    mfg_date=bc.opening_mfg_date,
                    expiry_date=bc.opening_expiry_date,
                )
                seeded_any = True
        if not seeded_any:
            # Listed branches exist at zero stock; legacy top-level opening still
            # supported when every branch row has opening_stock = 0.
            if data.opening_stock and data.opening_stock > 0:
                await enforce_branch_access(data.branch_id, user=user, db=db)
                await _seed_opening_stock(
                    db,
                    item=item,
                    branch_id=data.branch_id,
                    qty=int(data.opening_stock),
                    cost_price=float(data.cost_price),
                    batch_number=data.opening_batch_number,
                    mfg_date=data.opening_mfg_date,
                    expiry_date=data.opening_expiry_date,
                )
    else:
        await enforce_branch_access(data.branch_id, user=user, db=db)
        await _upsert_branch_config(
            db,
            item_id=item.id,
            branch_id=data.branch_id,
            is_available=True,
            cost_price=None,
            selling_price=None,
            reorder_level=None,
        )
        await _ensure_stock_row(db, item_id=item.id, branch_id=data.branch_id)
        if data.opening_stock and data.opening_stock > 0:
            await _seed_opening_stock(
                db,
                item=item,
                branch_id=data.branch_id,
                qty=int(data.opening_stock),
                cost_price=float(data.cost_price),
                batch_number=data.opening_batch_number,
                mfg_date=data.opening_mfg_date,
                expiry_date=data.opening_expiry_date,
            )

    # If the caller can directly commit (has approve permission), mark
    # the item approved now and record approver metadata.
    if direct:
        item.active = True
        item.status = ItemApprovalStatus.approved
        item.approval_status = ItemApprovalStatus.approved
        item.approved_by = user.name
        item.approved_at = datetime.utcnow()

    _log_item_history(
        db,
        user=user,
        item=item,
        event_type="item_created",
        action="Item Created",
        detail=f"Created item {item.name} ({item.sku})",
        metadata={"sku": item.sku, "name": item.name},
    )
    if not direct:
        from src.notifications.store import emit_item_pending, notify_refresh

        await emit_item_pending(db, item)
    await db.commit()
    status = item.approval_status.value if hasattr(item.approval_status, "value") else "approved"
    await _write_post_commit_audit(
        db,
        action="New Item Added",
        reference_id=item.sku or item.id,
        detail=(f"Created and approved item {item.name} ({item.sku})" if direct else f"Created item {item.name} ({item.sku})"),
        user=user,
        request=request,
        branch_id=data.branch_id,
        metadata={"item_id": item.id, "sku": item.sku, "name": item.name, "status": status},
    )
    if not direct:
        await notify_refresh()
    return {
        "id": item.id,
        "message": ("Item created and approved" if direct else "Item submitted for approval"),
        "approval_status": status,
        "status": status,
        "active": bool(item.active),
    }


@router.post("/import", dependencies=[Depends(require_perm("item_master.create"))])
async def import_items(
    file: UploadFile = File(...),
    request: Request = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    """Bulk import items from an Excel file. Returns a report with successes and row-level errors."""
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(400, detail=str(e))

    try:
        from io import BytesIO
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, detail=f"Failed to read Excel file: {e}")

    # Read header row and map column names to indices
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(400, detail="Spreadsheet must have a header row and at least one data row")
    headers = [str(h).strip().lower() if h is not None else None for h in rows[0]]

    # Simple mapping from common header names to model fields
    map_keys = {
        'name': 'name',
        'item name': 'name',
        'sku': 'sku',
        'barcode': 'barcode',
        'category': 'category_id',
        'brand': 'brand',
        'unit': 'unit',
        'cost': 'cost_price',
        'cost price': 'cost_price',
        'selling price': 'selling_price',
        'price': 'selling_price',
        'tax': 'tax_rate',
        'tax rate': 'tax_rate',
        'reorder': 'reorder_level',
        'reorder level': 'reorder_level',
        'batch tracking': 'batch_tracking',
        'expiry tracking': 'expiry_tracking',
        'active': 'active',
    }

    created = 0
    errors = []
    # Process rows starting from second row
    for idx, row in enumerate(rows[1:], start=2):
        try:
            data = {}
            branch_cells = {}
            for col_idx, cell in enumerate(row):
                key = headers[col_idx] if col_idx < len(headers) else None
                if not key:
                    continue
                mapped = map_keys.get(key)
                if not mapped:
                    # detect per-branch columns like "Opening Stock - Branch Name",
                    # "Opening Batch Number - Branch Name", "Cost Price - Branch Name",
                    # "Selling Price - Branch Name"
                    m = re.match(r'^(opening stock|opening batch number|cost price|selling price)\s*-\s*(.+)$', key)
                    if m:
                        col = m.group(1).strip()
                        branch_label = m.group(2).strip()
                        bmap = branch_cells.setdefault(branch_label, {})
                        bmap[col] = cell
                        continue
                    continue
                val = cell
                if mapped in {'cost_price', 'selling_price', 'tax_rate'} and val is not None:
                    try:
                        val = float(val)
                    except Exception:
                        raise ValueError(f"Invalid numeric value for {key}: {val}")
                if mapped in {'reorder_level'} and val is not None:
                    try:
                        val = int(val)
                    except Exception:
                        raise ValueError(f"Invalid integer for {key}: {val}")
                if mapped in {'batch_tracking', 'expiry_tracking', 'active'} and val is not None:
                    sval = str(val).strip().lower()
                    val = sval in {'1', 'true', 'yes', 'y'}
                data[mapped] = val

            if 'name' not in data or not data['name']:
                raise ValueError('Name is required')

            # Resolve category value -> id if provided. Try id first, then name.
            category_id = None
            if data.get('category_id'):
                cat_val = str(data.get('category_id')).strip()
                if cat_val:
                    # Try matching by id first
                    res = await db.execute(select(Category).where(Category.id == cat_val))
                    cat = res.scalar_one_or_none()
                    if not cat:
                        # Fallback to case-insensitive name match
                        res = await db.execute(select(Category).where(func.lower(Category.name) == cat_val.lower()))
                        cat = res.scalar_one_or_none()
                    if cat:
                        category_id = cat.id
                    else:
                        # Create the category automatically and use it
                        new_cat = Category(id=f"cat-{uuid.uuid4().hex[:8]}", name=cat_val, icon="📦")
                        db.add(new_cat)
                        await db.flush()
                        category_id = new_cat.id
            # Determine direct approval
            direct = await can_direct_commit(user, db, 'item_master.approve')
            initial_status = ItemApprovalStatus.approved if direct else ItemApprovalStatus.pending

            item = Item(
                id=str(uuid.uuid4()),
                name=str(data.get('name')).strip(),
                sku=(data.get('sku') or f"SKU-{uuid.uuid4().hex[:6].upper()}"),
                barcode=data.get('barcode'),
                country_of_origin=None,
                category_id=category_id,
                brand=data.get('brand'),
                unit=data.get('unit') or 'Pcs',
                cost_price=float(data.get('cost_price') or 0),
                selling_price=float(data.get('selling_price') or 0),
                wholesale_discount_pct=0,
                staff_discount_pct=0,
                tax_rate=float(data.get('tax_rate') or 0),
                hsn_code=None,
                reorder_level=int(data.get('reorder_level') or 0),
                is_packaging=False,
                packaging_quantity=None,
                emoji='📦',
                batch_tracking=bool(data.get('batch_tracking') or False),
                expiry_tracking=bool(data.get('expiry_tracking') or False),
                active=bool(direct),
                approval_status=initial_status,
                status=initial_status,
                created_by=user.name,
            )
            db.add(item)
            await db.flush()
            # Process per-branch columns (if any)
            if branch_cells:
                for branch_label, cols in branch_cells.items():
                    try:
                        # find branch by name (case-insensitive) or id
                        branch_res = await db.execute(select(Branch).where(func.lower(Branch.name) == branch_label.lower()))
                        branch = branch_res.scalar_one_or_none()
                        if branch is None:
                            # also try matching by id
                            branch_res = await db.execute(select(Branch).where(Branch.id == branch_label))
                            branch = branch_res.scalar_one_or_none()
                        if branch is None:
                            errors.append({'row': idx, 'error': f'Unknown branch: {branch_label}'})
                            continue
                        # ensure user may access this branch
                        try:
                            await enforce_branch_access(branch.id, user=user, db=db)
                        except Exception as e:
                            errors.append({'row': idx, 'error': f'No access to branch {branch_label}: {e}'})
                            continue

                        # upsert branch config with provided cost/selling price if present
                        branch_cost = None
                        branch_selling = None
                        if 'cost price' in cols and cols.get('cost price') is not None and cols.get('cost price') != '':
                            try:
                                branch_cost = float(cols.get('cost price'))
                            except Exception:
                                errors.append({'row': idx, 'error': f'Invalid cost price for branch {branch_label}: {cols.get("cost price")}'})
                                branch_cost = None
                        if 'selling price' in cols and cols.get('selling price') is not None and cols.get('selling price') != '':
                            try:
                                branch_selling = float(cols.get('selling price'))
                            except Exception:
                                errors.append({'row': idx, 'error': f'Invalid selling price for branch {branch_label}: {cols.get("selling price")}'})
                                branch_selling = None

                        await _upsert_branch_config(
                            db,
                            item_id=item.id,
                            branch_id=branch.id,
                            is_available=True,
                            cost_price=branch_cost,
                            selling_price=branch_selling,
                        )

                        # Seed opening stock if provided
                        if 'opening stock' in cols and cols.get('opening stock') not in (None, ''):
                            try:
                                qty = int(cols.get('opening stock'))
                            except Exception:
                                errors.append({'row': idx, 'error': f'Invalid opening stock for branch {branch_label}: {cols.get("opening stock")}'})
                                qty = 0
                            if qty > 0:
                                batch_number = cols.get('opening batch number') or None
                                # prefer branch-specific cost if provided otherwise use item's cost_price
                                cost_for_seed = branch_cost if branch_cost is not None else float(data.get('cost_price') or 0)
                                await _seed_opening_stock(
                                    db,
                                    item=item,
                                    branch_id=branch.id,
                                    qty=qty,
                                    cost_price=cost_for_seed,
                                    batch_number=batch_number,
                                )
                    except Exception as e:
                        errors.append({'row': idx, 'error': f'Failed processing branch {branch_label}: {e}'})
            # Do not seed opening stock from imports — operator can set per-branch later
            if direct:
                item.active = True
                item.status = ItemApprovalStatus.approved
                item.approval_status = ItemApprovalStatus.approved
                item.approved_by = user.name
                item.approved_at = datetime.utcnow()

            _log_item_history(
                db,
                user=user,
                item=item,
                event_type='item_created',
                action='Item Created (import)',
                detail=f'Imported item {item.name} ({item.sku})',
                metadata={'sku': item.sku, 'name': item.name},
            )
            created += 1
        except Exception as e:
            # If a DB error (e.g. IntegrityError) occurred during flush/insert,
            # the session transaction will be in rollback state. Roll back to
            # clear the session so we can continue processing remaining rows.
            try:
                await db.rollback()
            except Exception:
                pass

            # Provide a concise, user-friendly message for duplicate-SKU errors
            from sqlalchemy.exc import IntegrityError as SAIntegrityError
            orig = getattr(e, 'orig', None)
            msg = str(e)
            if isinstance(e, SAIntegrityError) or (orig is not None and 'duplicate key value violates unique constraint' in str(orig)):
                # Check for items_sku_key or mention of (sku)= in the DB error
                if (('items_sku_key' in msg) or (orig is not None and 'items_sku_key' in str(orig))) or ('Key (sku)' in msg) or (orig is not None and 'Key (sku)' in str(orig)):
                    sku_val = data.get('sku') if isinstance(data, dict) else None
                    sku_display = sku_val or '<unknown>'
                    errors.append({'row': idx, 'error': f"SKU already exists - {sku_display}"})
                    continue
            # Fallback: generic error string
            errors.append({'row': idx, 'error': msg})

    await db.commit()
    return {'created': created, 'errors': errors}


@router.get("/import/template", dependencies=[Depends(require_perm("item_master.create"))])
async def download_import_template(db: AsyncSession = Depends(get_db)):
    """Generate an Excel import template that includes current branches as per-branch columns."""
    res = await db.execute(select(Branch).order_by(asc(Branch.name)))
    branches = res.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    headers = [
        "Name",
        "SKU",
        "Barcode",
        "Category",
        "Brand",
        "Unit",
        "Cost Price",
        "Selling Price",
        "Tax Rate",
        "Reorder Level",
        "Batch Tracking",
        "Expiry Tracking",
        "Active",
    ]

    for b in branches:
        # Use branch name in header to be user-friendly
        label = b.name or b.id
        headers.append(f"Opening Stock - {label}")
        headers.append(f"Opening Batch Number - {label}")
        headers.append(f"Cost Price - {label}")
        headers.append(f"Selling Price - {label}")

    ws.append(headers)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': 'attachment; filename="item_import_template.xlsx"'
    })



@router.get("/{item_id}", dependencies=[Depends(require_perm(*ITEM_CATALOG_READ))])
async def get_item(item_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Item)
        .where(Item.id == item_id)
        .options(selectinload(Item.category))
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    return {
        "id": item.id,
        "name": item.name,
        "sku": item.sku,
        "barcode": item.barcode,
        "country_of_origin": item.country_of_origin,
        "categoryId": item.category_id,
        "categoryName": item.category.name if item.category else "Uncategorized",
        "brand": item.brand,
        "unit": item.unit,
        "default_cost_price": item.cost_price,
        "cost_price": item.cost_price,
        "default_selling_price": item.selling_price,
        "selling_price": item.selling_price,
        "wholesale_discount_pct": float(getattr(item, "wholesale_discount_pct", 0) or 0),
        "staff_discount_pct": float(getattr(item, "staff_discount_pct", 0) or 0),
        "tax_rate": item.tax_rate,
        "hsn_code": item.hsn_code,
        "default_reorder_level": item.reorder_level,
        "reorder_level": item.reorder_level,
        "is_packaging": item.is_packaging,
        "packaging_quantity": item.packaging_quantity,
        "emoji": item.emoji,
        "batch_tracking": item.batch_tracking,
        "expiry_tracking": item.expiry_tracking,
        "active": item.active,
        "approval_status": _resolve_item_status_value(item),
        "status": _resolve_item_status_value(item),
        "created_by": item.created_by,
        "approved_by": item.approved_by,
        "rejected_by": item.rejected_by,
        "rejection_reason": item.rejection_reason,
    }

@router.post("/{item_id}/approve", dependencies=[Depends(require_perm("item_master.approve"))])
async def approve_item(
    item_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    result = await db.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    # Extra guard: ensure the resolved permission set actually grants approve.
    # This double-check prevents accidental bypasses caused by mismatched
    # role mappings or demo-mode fallbacks.
    if not await user_can(user, db, "item_master.approve"):
        raise HTTPException(403, "Missing permission: item_master.approve")
    current_status = _resolve_item_status_value(item)
    if item.created_by and item.created_by == user.name and current_status in {ItemApprovalStatus.pending_approval.value, ItemApprovalStatus.pending.value}:
        raise HTTPException(403, "You cannot approve your own item submission")
    if current_status == ItemApprovalStatus.approved.value and item.active:
        return {"id": item.id, "approval_status": "approved", "already_processed": True}
    if current_status == ItemApprovalStatus.rejected.value:
        raise HTTPException(400, "Rejected items cannot be approved — create a new item")
    item.active = True
    item.status = ItemApprovalStatus.approved
    item.approval_status = ItemApprovalStatus.approved
    item.approved_by = user.name
    item.approved_at = datetime.utcnow()
    item.rejected_by = None
    item.rejected_at = None
    item.rejection_reason = None
    from src.notifications.store import notify_refresh, resolve_notification

    await resolve_notification(db, f"approval.item_master_pending:{item.id}")
    await db.commit()
    await _write_post_commit_audit(
        db,
        action="Item Approved",
        reference_id=item.sku or item.id,
        detail=f"Approved item {item.name} ({item.sku})",
        user=user,
        request=request,
        branch_id=None,
        metadata={"item_id": item.id, "sku": item.sku, "name": item.name},
    )
    _log_item_history(
        db,
        user=user,
        item=item,
        event_type="item_approved",
        action="Item Approved",
        detail=f"Approved item {item.name} ({item.sku})",
        metadata={"item_id": item.id, "sku": item.sku, "name": item.name},
    )
    await db.commit()
    await notify_refresh()
    return {"id": item.id, "approval_status": "approved", "status": "approved", "active": True}


class ItemReject(BaseModel):
    notes: Optional[str] = None
    reason: Optional[str] = None


@router.post("/{item_id}/reject", dependencies=[Depends(require_perm("item_master.approve"))])
async def reject_item(
    item_id: str,
    body: ItemReject,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    result = await db.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    if not await user_can(user, db, "item_master.approve"):
        raise HTTPException(403, "Missing permission: item_master.approve")
    current_status = _resolve_item_status_value(item)
    if current_status == ItemApprovalStatus.rejected.value:
        return {"id": item.id, "approval_status": "rejected", "status": "rejected", "already_processed": True}
    if current_status not in {ItemApprovalStatus.pending.value, ItemApprovalStatus.pending_approval.value}:
        raise HTTPException(400, "Only pending items can be rejected")
    reason = (body.reason or body.notes or "").strip()
    if not reason:
        raise HTTPException(400, "Rejection reason is required")
    item.status = ItemApprovalStatus.rejected
    item.approval_status = ItemApprovalStatus.rejected
    item.active = False
    item.approved_by = None
    item.rejected_by = user.name
    item.rejected_at = datetime.utcnow()
    item.rejection_reason = reason
    item.approved_at = None
    from src.notifications.store import notify_refresh, resolve_notification

    await resolve_notification(db, f"approval.item_master_pending:{item.id}")
    await db.commit()
    await _write_post_commit_audit(
        db,
        action="Item Rejected",
        reference_id=item.sku or item.id,
        detail=f"Rejected item {item.name} ({item.sku})",
        user=user,
        request=request,
        branch_id=None,
        metadata={"item_id": item.id, "sku": item.sku, "name": item.name, "reason": reason},
    )
    _log_item_history(
        db,
        user=user,
        item=item,
        event_type="item_rejected",
        action="Item Rejected",
        detail=f"Rejected item {item.name} ({item.sku})",
        metadata={"item_id": item.id, "sku": item.sku, "name": item.name, "reason": reason},
    )
    await db.commit()
    await notify_refresh()
    return {"id": item.id, "approval_status": "rejected", "status": "rejected", "reason": reason}

@router.put("/{item_id}", dependencies=[Depends(require_perm("item_master.edit"))])
async def update_item(
    item_id: str,
    data: ItemCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    result = await db.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    await _validate_category_id(data.category_id, db)
    if data.is_packaging and (data.packaging_quantity is None or data.packaging_quantity <= 0):
        raise HTTPException(400, "Packaging quantity must be greater than zero when packaging is enabled")

    was_tracked = bool(item.batch_tracking)
    before = {
        "name": item.name,
        "sku": item.sku,
        "barcode": item.barcode,
        "country_of_origin": item.country_of_origin,
        "cost_price": item.cost_price,
        "selling_price": item.selling_price,
        "wholesale_discount_pct": getattr(item, "wholesale_discount_pct", 0) or 0,
        "staff_discount_pct": getattr(item, "staff_discount_pct", 0) or 0,
        "tax_rate": item.tax_rate,
        "reorder_level": item.reorder_level,
        "batch_tracking": item.batch_tracking,
        "expiry_tracking": item.expiry_tracking,
    }
    item.name = data.name
    item.sku = data.sku or item.sku
    item.barcode = data.barcode
    item.country_of_origin = data.country_of_origin
    item.category_id = data.category_id or None
    item.brand = data.brand
    item.unit = data.unit
    item.cost_price = data.cost_price
    item.selling_price = data.selling_price
    item.wholesale_discount_pct = float(data.wholesale_discount_pct or 0)
    item.staff_discount_pct = float(data.staff_discount_pct or 0)
    item.tax_rate = data.tax_rate
    item.hsn_code = data.hsn_code
    item.reorder_level = data.reorder_level
    item.is_packaging = data.is_packaging
    item.packaging_quantity = data.packaging_quantity if data.is_packaging else None
    item.batch_tracking = data.batch_tracking
    item.expiry_tracking = data.expiry_tracking if data.batch_tracking else False

    toggle_meta = await _apply_batch_tracking_toggle(
        db,
        item,
        previously_tracked=was_tracked,
        enable=bool(data.batch_tracking),
    )

    _log_item_history(
        db,
        user=user,
        item=item,
        event_type="item_updated",
        action="Item Updated",
        detail=f"Updated item {item.name} ({item.sku})",
        metadata={
            "sku": item.sku,
            "name": item.name,
            "updated_fields": [
                key for key in before.keys() if before.get(key) != getattr(item, key)
            ],
            "changes": {
                key: {"before": before.get(key), "after": getattr(item, key)}
                for key in before.keys()
                if before.get(key) != getattr(item, key)
            },
        },
    )
    await _write_post_commit_audit(
        db,
        action="Item Updated",
        reference_id=item.sku or item.id,
        detail=f"Updated item {item.name} ({item.sku})",
        user=user,
        request=request,
        branch_id=data.branch_id,
        metadata={"item_id": item.id, "sku": item.sku, "name": item.name},
    )

    await db.commit()
    await db.refresh(item)
    out: dict = {"id": item.id, "message": "Item updated"}
    if toggle_meta:
        out["batch_tracking_change"] = toggle_meta
    return out

@router.get(
    "/{item_id}/branches/{branch_id}/batch-info",
    dependencies=[Depends(require_perm("items.view"))],
)
async def get_branch_batch_info(
    item_id: str,
    branch_id: str,
    db: AsyncSession = Depends(get_db),
):
    """Check if a branch has batches or stock before removal.
    
    Used by the Item Master form to warn users before deleting a branch
    that has existing batches/stock.
    """
    item_res = await db.execute(select(Item).where(Item.id == item_id))
    item = item_res.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    
    info = await _get_branch_batch_info(db, item_id, branch_id)
    return info

@router.get(
    "/{item_id}/branches",
    dependencies=[Depends(require_perm("items.view", "item_master.view"))],
)
async def get_item_branches(item_id: str, db: AsyncSession = Depends(get_db)):
    item_res = await db.execute(select(Item).where(Item.id == item_id))
    item = item_res.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    branches = (await db.execute(
        select(Branch).where(Branch.active == True).order_by(Branch.name)  # noqa: E712
    )).scalars().all()
    cfg_res = await db.execute(
        select(ItemBranchConfig).where(ItemBranchConfig.item_id == item_id)
    )
    cfg_map = {c.branch_id: c for c in cfg_res.scalars().all()}
    stock_res = await db.execute(
        select(ItemStock).where(ItemStock.item_id == item_id)
    )
    stock_map = {s.branch_id: int(s.quantity or 0) for s in stock_res.scalars().all()}

    out = []
    for br in branches:
        cfg = cfg_map.get(br.id)
        out.append({
            "branch_id": br.id,
            "branch_name": br.name,
            "branch_code": br.code,
            "is_available": bool(cfg.is_available) if cfg else False,
            "cost_price": cfg.cost_price if cfg else None,
            "selling_price": cfg.selling_price if cfg else None,
            "reorder_level": cfg.reorder_level if cfg else None,
            "effective_cost_price": effective_cost_price(
                item.cost_price,
                cfg.cost_price if cfg else None,
            ),
            "effective_selling_price": effective_selling_price(
                item.selling_price,
                cfg.selling_price if cfg else None,
            ),
            "effective_reorder_level": effective_reorder_level(
                item.reorder_level,
                cfg.reorder_level if cfg else None,
            ),
            "available_stock": stock_map.get(br.id, 0),
        })
    return {
        "item_id": item.id,
        "default_cost_price": item.cost_price,
        "default_selling_price": item.selling_price,
        "default_reorder_level": item.reorder_level,
        "branches": out,
    }


@router.put(
    "/{item_id}/branches",
    dependencies=[Depends(require_perm("item_master.edit"))],
)
async def update_item_branches(
    item_id: str,
    data: BranchConfigBulk,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    item_res = await db.execute(select(Item).where(Item.id == item_id))
    item = item_res.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    branch_ids = set((await db.execute(select(Branch.id))).scalars().all())
    any_branch_change = False
    total_batches_deleted = 0
    total_stock_deleted = 0
    removed_branches_info = []
    for bc in data.branches:
        if bc.branch_id not in branch_ids:
            raise HTTPException(400, f"Unknown branch: {bc.branch_id}")

        existing_cfg = (
            await db.execute(
                select(ItemBranchConfig).where(
                    ItemBranchConfig.item_id == item_id,
                    ItemBranchConfig.branch_id == bc.branch_id,
                )
            )
        ).scalar_one_or_none()

        prev_is_available = bool(existing_cfg.is_available) if existing_cfg else False
        prev_cost = existing_cfg.cost_price if existing_cfg else None
        prev_sell = existing_cfg.selling_price if existing_cfg else None
        prev_reorder = existing_cfg.reorder_level if existing_cfg else None

        branch_changed = (
            existing_cfg is None
            or prev_is_available != bool(bc.is_available)
            or prev_cost != bc.cost_price
            or prev_sell != bc.selling_price
            or prev_reorder != bc.reorder_level
        )

        await _upsert_branch_config(
            db,
            item_id=item_id,
            branch_id=bc.branch_id,
            is_available=bc.is_available,
            cost_price=bc.cost_price,
            selling_price=bc.selling_price,
            reorder_level=bc.reorder_level,
        )
        if bc.is_available:
            await _ensure_stock_row(db, item_id=item_id, branch_id=bc.branch_id)
        if bc.is_available and int(bc.opening_stock or 0) > 0:
            branch_changed = True
            branch_cost = effective_cost_price(item.cost_price, bc.cost_price)
            await _seed_opening_stock(
                db,
                item=item,
                branch_id=bc.branch_id,
                qty=int(bc.opening_stock),
                cost_price=branch_cost,
                batch_number=bc.opening_batch_number,
                mfg_date=bc.opening_mfg_date,
                expiry_date=bc.opening_expiry_date,
            )
        
        # When a branch is being removed (is_available = False), delete its batches and stock
        if not bc.is_available:
            # Delete batches for this branch
            batch_delete_res = await db.execute(
                delete(ItemBatch).where(
                    ItemBatch.item_id == item_id,
                    ItemBatch.branch_id == bc.branch_id,
                )
            )
            batches_deleted = int(batch_delete_res.rowcount or 0)
            
            # Delete stock for this branch
            stock_delete_res = await db.execute(
                delete(ItemStock).where(
                    ItemStock.item_id == item_id,
                    ItemStock.branch_id == bc.branch_id,
                )
            )
            stock_deleted = int(stock_delete_res.rowcount or 0)
            
            total_batches_deleted += batches_deleted
            total_stock_deleted += stock_deleted
            
            if batches_deleted > 0 or stock_deleted > 0:
                branch_changed = True
                # Get branch name for logging
                branch_res = await db.execute(select(Branch.name).where(Branch.id == bc.branch_id))
                branch_name = branch_res.scalar_one_or_none() or bc.branch_id
                
                removed_branches_info.append({
                    "branch_id": bc.branch_id,
                    "branch_name": branch_name,
                    "batches_deleted": batches_deleted,
                    "stock_deleted": stock_deleted,
                })
                
                _log_item_history(
                    db,
                    user=user,
                    item=item,
                    event_type="item_branch_removed",
                    action="Item Branch Removed",
                    detail=f"Removed branch {branch_name} from item {item.name} ({item.sku}). Deleted {batches_deleted} batch(es) and {stock_deleted} stock row(s).",
                    metadata={
                        "branch_id": bc.branch_id,
                        "branch_name": branch_name,
                        "batches_deleted": batches_deleted,
                        "stock_deleted": stock_deleted,
                        "sku": item.sku,
                        "name": item.name,
                    },
                    risk="medium",
                )

        if branch_changed:
            any_branch_change = True

    if any_branch_change:
        _log_item_history(
            db,
            user=user,
            item=item,
            event_type="item_branch_config_updated",
            action="Item Branch Config Updated",
            detail=f"Updated branch configuration for item {item.name} ({item.sku})",
            metadata={
                "branch_count": len(data.branches),
                "sku": item.sku,
                "name": item.name,
            },
        )
    await db.commit()
    response = {"message": "Branch configuration updated"}
    if total_batches_deleted > 0 or total_stock_deleted > 0:
        response["branches_removed"] = removed_branches_info
        response["total_batches_deleted"] = total_batches_deleted
        response["total_stock_deleted"] = total_stock_deleted
    
    return response



@router.patch("/{item_id}", dependencies=[Depends(require_perm("item_master.edit"))])
async def patch_item(
    item_id: str,
    data: ItemPatch,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    result = await db.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    was_tracked = bool(item.batch_tracking)
    updates = data.model_dump(exclude_unset=True)
    branch_id = updates.pop("branch_id", None)
    before = {
        key: getattr(item, key)
        for key in [
            "name",
            "sku",
            "barcode",
            "country_of_origin",
            "cost_price",
            "selling_price",
            "wholesale_discount_pct",
            "staff_discount_pct",
            "is_packaging",
            "packaging_quantity",
            "tax_rate",
            "reorder_level",
            "batch_tracking",
            "expiry_tracking",
            "active",
            "category_id",
            "brand",
            "unit",
            "hsn_code",
        ]
        if hasattr(item, key)
    }
    toggle_meta = None
    if "active" in updates and updates["active"] is False:
        blockers = await _item_inventory_blockers(db, item_id, branch_id=branch_id)
        if blockers:
            raise HTTPException(
                400,
                "Cannot deactivate item: " + "; ".join(blockers),
            )
    if "batch_tracking" in updates:
        will_track = bool(updates["batch_tracking"])
        toggle_meta = await _apply_batch_tracking_toggle(
            db,
            item,
            previously_tracked=was_tracked,
            enable=will_track,
        )
        if not will_track:
            updates["expiry_tracking"] = False
    for k, v in updates.items():
        setattr(item, k, v)

    changes = {
        key: {"before": before.get(key), "after": getattr(item, key)}
        for key in updates.keys()
        if before.get(key) != getattr(item, key)
    }

    _log_item_history(
        db,
        user=user,
        item=item,
        event_type="item_updated",
        action="Item Updated",
        detail=f"Updated item {item.name} ({item.sku})",
        metadata={
            "updated_fields": list(updates.keys()),
            "changes": changes,
            "sku": item.sku,
            "name": item.name,
        },
    )
    await db.commit()
    out: dict = {"message": "Updated"}
    if toggle_meta:
        out["batch_tracking_change"] = toggle_meta
    return out

@router.delete("/{item_id}", dependencies=[Depends(require_perm("item_master.delete"))])
async def delete_item(
    item_id: str,
    branch_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    blockers = await _item_inventory_blockers(db, item_id, branch_id=branch_id)
    if blockers:
        raise HTTPException(
            400,
            "Cannot delete item: " + "; ".join(blockers),
        )

    if await _item_has_historical_references(db, item_id, branch_id=branch_id):
        raise HTTPException(
            400,
            "Cannot delete item because it has historical transactional references",
        )

    if branch_id:
        adjustment_ref = await db.execute(
            select(AdjustmentRequest.id)
            .where(
                AdjustmentRequest.item_id == item_id,
                AdjustmentRequest.branch_id == branch_id,
            )
            .limit(1)
        )
        if adjustment_ref.scalar_one_or_none():
            raise HTTPException(
                400,
                "Cannot delete item because it is still referenced by adjustment requests in this branch",
            )

        stock_adjustment_ref = await db.execute(
            select(StockAdjustment.id)
            .where(
                StockAdjustment.item_id == item_id,
                StockAdjustment.branch_id == branch_id,
            )
            .limit(1)
        )
        if stock_adjustment_ref.scalar_one_or_none():
            raise HTTPException(
                400,
                "Cannot delete item because it is still referenced by stock adjustment history in this branch",
            )

        await db.execute(
            delete(ItemBatch).where(
                ItemBatch.item_id == item_id,
                ItemBatch.branch_id == branch_id,
            )
        )
        await db.execute(
            delete(ItemBranchConfig).where(
                ItemBranchConfig.item_id == item_id,
                ItemBranchConfig.branch_id == branch_id,
            )
        )
        await db.execute(
            delete(ItemStock).where(
                ItemStock.item_id == item_id,
                ItemStock.branch_id == branch_id,
            )
        )

        # If the item no longer has any branch listings, delete the master
        # item row too when it is safely removable.
        remaining_branches = await db.execute(
            select(func.count(ItemBranchConfig.id)).where(ItemBranchConfig.item_id == item_id)
        )
        if int(remaining_branches.scalar() or 0) == 0:
            if not await _item_has_historical_references(db, item_id):
                await db.execute(delete(Item).where(Item.id == item_id))

        await db.commit()
        return {"deleted": item_id, "branch_id": branch_id}

    adjustment_ref = await db.execute(
        select(AdjustmentRequest.id).where(AdjustmentRequest.item_id == item_id).limit(1)
    )
    if adjustment_ref.scalar_one_or_none():
        raise HTTPException(
            400,
            "Cannot delete item because it is still referenced by adjustment requests",
        )

    stock_adjustment_ref = await db.execute(
        select(StockAdjustment.id).where(StockAdjustment.item_id == item_id).limit(1)
    )
    if stock_adjustment_ref.scalar_one_or_none():
        raise HTTPException(
            400,
            "Cannot delete item because it is still referenced by stock adjustment history",
        )

    await db.execute(delete(ItemBatch).where(ItemBatch.item_id == item_id))
    await db.execute(delete(ItemBranchConfig).where(ItemBranchConfig.item_id == item_id))
    await db.execute(delete(ItemStock).where(ItemStock.item_id == item_id))
    try:
        await db.execute(delete(Item).where(Item.id == item_id))
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            400,
            "Cannot delete item because it is still referenced by other records",
        )
    await db.commit()
    return {"deleted": item_id}

@router.post("/adjust", dependencies=[Depends(require_perm("adjustments.approve"))])
async def adjust_stock(data: StockAdjustRequest, db: AsyncSession = Depends(get_db)):
    """Direct stock apply — restricted to approvers (e.g. API integrations).

    UI flows should use POST /adjustments/ and approve instead.
    """
    from src.routes._stock_adjust_apply import apply_stock_adjustment

    try:
        result = await apply_stock_adjustment(
            db,
            item_id=data.item_id,
            branch_id=data.branch_id,
            new_qty=int(data.new_qty),
            reason=data.reason,
            notes=data.notes,
            batch_id=data.batch_id,
        )
    except LookupError:
        raise HTTPException(404, "Batch not found")
    except ValueError as e:
        raise HTTPException(400, str(e))
    await db.commit()
    return result


# ─── Batches (per item) ──────────────────────────────────────────────────────
def _batch_dict(b: ItemBatch, *, item_name: Optional[str] = None) -> dict:
    return {
        "id":            b.id,
        "itemId":        b.item_id,
        "itemName":      item_name,
        "branchId":      b.branch_id,
        "batchNumber":   b.batch_number,
        "mfgDate":       b.mfg_date,
        "expiryDate":    b.expiry_date,
        "quantity":      int(b.quantity or 0),
        "initialQty":    int(b.initial_qty or 0),
        "costPrice":     float(b.cost_price or 0),
        "vendorId":      b.vendor_id,
        "sourceType":    b.source_type,
        "sourceRef":     b.source_ref,
        "receivedDate":  b.received_date,
        "notes":         b.notes,
        "active":        bool(b.active),
        "createdAt":     b.created_at.isoformat() if b.created_at else None,
        "updatedAt":     b.updated_at.isoformat() if b.updated_at else None,
    }


@router.get(
    "/{item_id}/batches",
    dependencies=[Depends(require_perm(*ITEM_CATALOG_READ))],
)
async def list_item_batches(
    item_id: str,
    branch_id: Optional[str] = Depends(enforce_branch_access_optional),
    include_empty: bool = False,
    include_inactive: bool = False,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    """List batches for an item. Default sort is FEFO-friendly (nearest expiry
    first, NULL expiries last, ties broken by received_date) which is also a
    sensible FIFO-ish order for the UI.

    Inactive (quarantined) batches are omitted by default so POS / transfer
    allocation cannot pick them. Pass ``include_inactive=true`` from the batch
    management UI when the operator needs to review or reactivate lots."""
    item_res = await db.execute(select(Item).where(Item.id == item_id))
    item = item_res.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    if branch_id is None and not getattr(user, "all_branches", False):
        branch_ids = await get_allowed_branch_ids(user, db)
        if not branch_ids:
            return []
        branch_id = branch_ids[0]

    q = select(ItemBatch).where(ItemBatch.item_id == item_id)
    if branch_id:
        q = q.where(ItemBatch.branch_id == branch_id)
    if not include_empty:
        q = q.where(ItemBatch.quantity > 0)
    if not include_inactive:
        q = q.where(ItemBatch.active == True)  # noqa: E712
    # Same ordering as _atomic._batch_order("fefo") — keep them in sync.
    q = q.order_by(
        text("(expiry_date IS NULL)"),
        asc(ItemBatch.expiry_date),
        asc(ItemBatch.received_date),
        asc(ItemBatch.created_at),
    )
    res = await db.execute(q)
    rows = [_batch_dict(b, item_name=item.name) for b in res.scalars().all()]
    return {
        "item": {
            "id":              item.id,
            "name":            item.name,
            "sku":             item.sku,
            "emoji":           item.emoji,
            "batch_tracking":  bool(item.batch_tracking),
            "expiry_tracking": bool(item.expiry_tracking),
        },
        "items": rows,
        "total": len(rows),
    }


@router.post(
    "/{item_id}/batches",
    status_code=201,
    dependencies=[Depends(require_perm("items.adjust"))],
)
async def create_item_batch(
    item_id: str,
    data: BatchCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    """Manually add a batch (e.g. correcting historical receipts, opening new
    stock for an existing item). Flips `batch_tracking` on as a side-effect
    if it wasn't already — once you have a batch, you're tracked."""
    item_res = await db.execute(select(Item).where(Item.id == item_id))
    item = item_res.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    if data.qty <= 0:
        raise HTTPException(400, "qty must be > 0")

    await enforce_branch_access_optional(data.branch_id, user=user, db=db)

    _raise_batch_date_errors(validate_batch_dates(
        mfg_date=data.mfg_date,
        expiry_date=data.expiry_date,
        received_date=data.received_date,
        require_expiry=bool(item.expiry_tracking),
    ))

    try:
        batch = await add_batch_atomic(
            db,
            item_id=item_id,
            branch_id=data.branch_id,
            qty=int(data.qty),
            batch_number=data.batch_number,
            mfg_date=data.mfg_date,
            expiry_date=data.expiry_date,
            cost_price=float(data.cost_price or item.cost_price or 0),
            vendor_id=data.vendor_id,
            source_type="manual",
            received_date=data.received_date,
            notes=data.notes,
        )
    except ValueError as e:
        # Duplicate batch number for this item/branch (see add_batch_atomic).
        raise HTTPException(400, str(e))
    await db.commit()
    return _batch_dict(batch, item_name=item.name)


@router.patch(
    "/batches/{batch_id}",
    dependencies=[Depends(require_perm("items.adjust"))],
)
async def patch_batch(
    batch_id: str,
    data: BatchPatch,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    """Edit batch metadata (dates, notes, number, active flag). Quantity
    changes must go through the stock-adjust endpoint so item_stock stays
    in sync."""
    res = await db.execute(select(ItemBatch).where(ItemBatch.id == batch_id))
    b = res.scalar_one_or_none()
    if not b:
        raise HTTPException(404, "Batch not found")
    await enforce_branch_access_optional(b.branch_id, user=user, db=db)
    updates = data.model_dump(exclude_unset=True)
    if any(k in updates for k in ("mfg_date", "expiry_date")):
        exp_flag = await db.execute(
            select(Item.expiry_tracking).where(Item.id == b.item_id)
        )
        require_exp = bool(exp_flag.scalar_one_or_none())

        def _norm_date(v: Optional[str]) -> Optional[str]:
            s = (v or "").strip()
            return s or None

        new_exp = _norm_date(
            updates["expiry_date"] if "expiry_date" in updates else b.expiry_date
        )
        # UI often re-submits the existing expiry on mfg-only edits — compare
        # values, not just key presence, so already-expired lots stay editable.
        expiry_changing = (
            "expiry_date" in updates
            and _norm_date(updates.get("expiry_date")) != _norm_date(b.expiry_date)
        )
        _raise_batch_date_errors(validate_batch_dates(
            mfg_date=updates.get("mfg_date", b.mfg_date),
            expiry_date=new_exp,
            require_expiry=require_exp and expiry_changing,
            allow_past_expiry=not expiry_changing,
        ))
    for k, v in updates.items():
        setattr(b, k, v)
    await db.commit()
    return _batch_dict(b)


# ─── Near-expiry rollup (used by Items page KPI) ─────────────────────────────
@router.get(
    "/batches/near-expiry",
    dependencies=[Depends(require_perm("items.view", "item_master.view"))],
)
async def near_expiry_batches(
    branch_id: Optional[str] = Depends(enforce_branch_access_optional),
    within_days: int = Query(30, ge=1, le=365),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    """List active batches whose expiry_date falls within the next N days
    (default 30). Drives the "Near Expiry" tab on the Items page so the
    operator can see which sellable lots to push out first."""
    from datetime import timedelta

    today_str   = datetime.utcnow().strftime("%Y-%m-%d")
    horizon_str = (datetime.utcnow() + timedelta(days=int(within_days))).strftime("%Y-%m-%d")

    q = (
        select(ItemBatch, Item)
        .join(Item, Item.id == ItemBatch.item_id)
        .where(
            ItemBatch.quantity > 0,
            ItemBatch.active == True,  # noqa: E712
            ItemBatch.expiry_date.isnot(None),
            ItemBatch.expiry_date <= horizon_str,
        )
    )
    if branch_id is None and not getattr(user, "all_branches", False):
        branch_ids = await get_allowed_branch_ids(user, db)
        if not branch_ids:
            return {"items": [], "total": 0, "within_days": within_days, "today": today_str}
        branch_id = branch_ids[0]
    if branch_id:
        q = q.where(ItemBatch.branch_id == branch_id)
    q = q.order_by(asc(ItemBatch.expiry_date), asc(ItemBatch.received_date))

    res = await db.execute(q)
    out = []
    for b, item in res.all():
        d = _batch_dict(b, item_name=item.name)
        d["itemEmoji"] = item.emoji
        d["itemSku"]   = item.sku
        d["expired"]   = bool(b.expiry_date and b.expiry_date < today_str)
        out.append(d)
    return {"items": out, "total": len(out), "within_days": within_days, "today": today_str}
