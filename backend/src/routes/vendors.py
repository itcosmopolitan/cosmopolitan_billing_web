import uuid
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.database import get_db
from src.models import User, Vendor, VendorCreditEntry
from src.pagination import normalize_limit, normalize_skip, paged, resolve_sort
from src.routes._serializers import serialize_vendor
from src.permissions import VENDOR_PICKER_READ
from src.security import current_user, require_perm

router = APIRouter()

class VendorCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    gstin: Optional[str] = None
    payment_terms: str = "30 days"

class VendorUpdate(BaseModel):
    name: Optional[str] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    gstin: Optional[str] = None
    payment_terms: Optional[str] = None

@router.get("/", dependencies=[Depends(require_perm(*VENDOR_PICKER_READ))])
async def list_vendors(
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "asc",
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    sk = normalize_skip(skip)
    lim = normalize_limit(limit)
    q = select(Vendor)
    cq = select(func.count(Vendor.id))
    conds = []
    if search:
        term = f"%{search}%"
        search_filter = or_(
            Vendor.name.ilike(term),
            Vendor.contact_person.ilike(term),
            Vendor.phone.ilike(term),
            Vendor.email.ilike(term),
        )
        q = q.where(search_filter)
        cq = cq.where(search_filter)
        conds.append(search_filter)
    total = int((await db.execute(cq)).scalar() or 0)
    outstanding_total = float(
        (
            await db.execute(
                select(func.coalesce(func.sum(Vendor.outstanding), 0)).where(
                    and_(*conds) if conds else True
                )
            )
        ).scalar()
        or 0
    )
    wb_filter = and_(Vendor.outstanding > 0, *conds) if conds else (Vendor.outstanding > 0)
    with_balance = int((await db.execute(select(func.count(Vendor.id)).where(wb_filter))).scalar() or 0)
    sort_expr = resolve_sort(
        sort_by,
        sort_order,
        {
            "name": Vendor.name,
            "contact_person": Vendor.contact_person,
            "phone": Vendor.phone,
            "email": Vendor.email,
            "payment_terms": Vendor.payment_terms,
            "outstanding": Vendor.outstanding,
            "total_purchases": Vendor.total_purchases,
            "created_at": Vendor.created_at,
        },
        default_key="name",
        default_order="asc",
    )
    result = await db.execute(q.order_by(sort_expr).offset(sk).limit(lim))
    items = [serialize_vendor(v) for v in result.scalars().all()]
    return paged(
        items,
        total,
        sk,
        lim,
        summary={"outstandingTotal": outstanding_total, "withBalanceCount": with_balance},
    )

@router.post("/import", dependencies=[Depends(require_perm("vendors.create"))])
async def import_vendors(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    """Bulk import vendors from an Excel file."""
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(400, detail=str(e))

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, detail=f"Failed to read Excel file: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(400, detail="Spreadsheet must have a header row and at least one data row")

    headers = [str(h).strip().lower() if h is not None else None for h in rows[0]]
    map_keys = {
        'name': 'name',
        'vendor name': 'name',
        'company': 'name',
        'company / vendor name': 'name',
        'contact person': 'contact_person',
        'contact': 'contact_person',
        'phone': 'phone',
        'email': 'email',
        'address': 'address',
        'gst reg no': 'gstin',
        'gst number': 'gstin',
        'gstin': 'gstin',
        'payment terms': 'payment_terms',
        'terms': 'payment_terms',
        'active': 'active',
    }

    created = 0
    errors = []

    def _as_text(value):
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    for idx, row in enumerate(rows[1:], start=2):
        try:
            data = {}
            for col_idx, cell in enumerate(row):
                key = headers[col_idx] if col_idx < len(headers) else None
                if not key:
                    continue
                mapped = map_keys.get(key)
                if not mapped:
                    continue
                val = cell
                if mapped == 'active' and val is not None:
                    sval = str(val).strip().lower()
                    val = sval in {'1', 'true', 'yes', 'y'}
                data[mapped] = val

            raw_name = _as_text(data.get('name'))
            if not raw_name:
                raise ValueError('Vendor name is required')

            vendor = Vendor(
                id=str(uuid.uuid4()),
                name=raw_name,
                contact_person=_as_text(data.get('contact_person')) or None,
                phone=_as_text(data.get('phone')) or None,
                email=_as_text(data.get('email')) or None,
                address=_as_text(data.get('address')) or None,
                gstin=_as_text(data.get('gstin')) or None,
                payment_terms=(_as_text(data.get('payment_terms')) or '30 days') or '30 days',
                active=bool(data.get('active', True)),
                created_at=None,
            )
            db.add(vendor)
            await db.flush()
            created += 1
        except Exception as e:
            try:
                await db.rollback()
            except Exception:
                pass
            errors.append({'row': idx, 'error': str(e)})

    await db.commit()
    return {'created': created, 'errors': errors}


@router.get("/import/template", dependencies=[Depends(require_perm("vendors.create"))])
async def download_vendor_import_template():
    """Generate an Excel template for bulk vendor imports."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendors"
    ws.append(["Vendor Name", "Contact Person", "Phone", "Email", "Address", "GST Reg No", "Payment Terms", "Active"])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="vendor_import_template.xlsx"'},
    )


@router.get("/{vendor_id}", dependencies=[Depends(require_perm(*VENDOR_PICKER_READ))])
async def get_vendor(vendor_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vendor).where(Vendor.id == vendor_id))
    v = result.scalar_one_or_none()
    if not v:
        raise HTTPException(404, "Vendor not found")
    return serialize_vendor(v)

@router.get("/{vendor_id}/credit-ledger", dependencies=[Depends(require_perm("vendors.view"))])
async def vendor_credit_ledger(
    vendor_id: str,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    """Append-only vendor advance / overpayment credit ledger."""
    exists = (
        await db.execute(select(Vendor.id).where(Vendor.id == vendor_id))
    ).scalar_one_or_none()
    if not exists:
        raise HTTPException(404, "Vendor not found")

    total = int(
        (
            await db.execute(
                select(func.count(VendorCreditEntry.id)).where(
                    VendorCreditEntry.vendor_id == vendor_id
                )
            )
        ).scalar()
        or 0
    )
    sk = normalize_skip(skip)
    lim = normalize_limit(limit)
    rows = (
        await db.execute(
            select(VendorCreditEntry)
            .where(VendorCreditEntry.vendor_id == vendor_id)
            .order_by(VendorCreditEntry.created_at.desc(), VendorCreditEntry.id.desc())
            .offset(sk)
            .limit(lim)
        )
    ).scalars().all()

    def _entry_type_label(t: str) -> str:
        labels = {
            "payment_debit": "Credit payment",
            "overpayment": "Overpayment credit",
            "return_void_revoke": "Return void (revoke)",
            "void_restore": "Payment void (restore)",
            "void_revoke": "Payment void (revoke)",
        }
        return labels.get(t, t.replace("_", " ").title())

    items = [{
        "id": e.id,
        "entryType": e.entry_type,
        "entryLabel": _entry_type_label(e.entry_type),
        "delta": round(float(e.delta or 0), 2),
        "balanceBefore": round(float(e.balance_before or 0), 2),
        "balanceAfter": round(float(e.balance_after or 0), 2),
        "sourceType": e.source_type,
        "sourceRef": e.source_ref,
        "sourceNumber": e.source_number,
        "notes": e.notes,
        "date": e.date,
        "createdBy": e.created_by,
        "createdAt": e.created_at.isoformat() if e.created_at else None,
    } for e in rows]
    return paged(items, total, sk, lim)

@router.post("/", status_code=201, dependencies=[Depends(require_perm("vendors.create"))])
async def create_vendor(data: VendorCreate, db: AsyncSession = Depends(get_db)):
    v = Vendor(id=str(uuid.uuid4()), name=data.name, contact_person=data.contact_person,
               phone=data.phone, email=data.email, address=data.address,
               gstin=data.gstin, payment_terms=data.payment_terms)
    db.add(v)
    await db.commit()
    await db.refresh(v)
    return serialize_vendor(v)

@router.put("/{vendor_id}", dependencies=[Depends(require_perm("vendors.edit"))])
async def update_vendor(vendor_id: str, data: VendorUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vendor).where(Vendor.id == vendor_id))
    v = result.scalar_one_or_none()
    if not v:
        raise HTTPException(404, "Vendor not found")
    for k, val in data.model_dump(exclude_unset=True).items():
        setattr(v, k, val)
    await db.commit()
    await db.refresh(v)
    return serialize_vendor(v)
