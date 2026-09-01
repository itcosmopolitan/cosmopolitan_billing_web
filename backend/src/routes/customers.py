import uuid
from io import BytesIO
from typing import Optional, Union

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, StrictInt
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.database import get_db
from src.models import Branch, Customer, CustomerCreditEntry, User
from src.pagination import normalize_limit, normalize_skip, paged, resolve_sort
from src.routes._serializers import _build_customer_code, serialize_customer
from src.permissions import CUSTOMER_PICKER_READ
from src.security import require_perm, current_user, enforce_branch_access
from src.models import User

router = APIRouter()

_VALID_CUSTOMER_TYPES = frozenset({"retail", "wholesale", "staff"})


class CustomerCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    gst_in: Optional[str] = None
    branch_id: str
    credit_limit: float = 0
    customer_type: str = "retail"
    key_account_manager: Optional[str] = None
    credit_terms: Optional[StrictInt] = Field(default=None, ge=0)
    street1: str
    street2: Optional[str] = None
    street3: Optional[str] = None
    city: str
    state_province: Optional[str] = None
    country: str
    postal_code: Optional[str] = None


class CustomerUpdate(BaseModel):
    """Typed update body — restricts client-writeable fields. `outstanding`
    and `total_purchases` are derived by sales/payment flows and must not be
    settable via PATCH.

    Field names mirror the public API contract (`gst_in`, `customer_type`)
    used by `CustomerCreate` and `serialize_customer`, NOT the underlying
    SQLAlchemy column names (`gstin`, `type`). Pydantic v2's default
    `extra="ignore"` would silently drop API-named keys otherwise, leaving
    those columns silently un-updated on a round-trip GET → mutate → PUT.
    The handler below maps API names → column names via `_FIELD_TO_COLUMN`.
    """
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    gst_in: Optional[str] = None
    branch_id: Optional[str] = None
    credit_limit: Optional[float] = None
    customer_type: Optional[str] = None
    key_account_manager: Optional[str] = None
    credit_terms: Optional[StrictInt] = Field(default=None, ge=0)
    street1: Optional[str] = None
    street2: Optional[str] = None
    street3: Optional[str] = None
    city: Optional[str] = None
    state_province: Optional[str] = None
    country: Optional[str] = None
    postal_code: Optional[str] = None
    notes: Optional[str] = None
    active: Optional[bool] = None


def _normalize_customer_type(customer_type: Optional[str]) -> str:
    value = (customer_type or "retail").strip().lower()
    if value not in _VALID_CUSTOMER_TYPES:
        raise HTTPException(
            400,
            f"Invalid customer_type '{customer_type}'. Expected one of: retail, wholesale, staff",
        )
    return value


# Map CustomerUpdate field names → Customer ORM column names where they
# differ. Single source of truth so PUT and POST can't drift again.
_FIELD_TO_COLUMN = {
    "gst_in": "gstin",
    "customer_type": "type",
}

@router.get("/", dependencies=[Depends(require_perm(*CUSTOMER_PICKER_READ))])
async def list_customers(
    search: Optional[str] = None,
    customer_type: Optional[str] = None,
    branch_id: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "asc",
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    sk = normalize_skip(skip)
    lim = normalize_limit(limit)
    q = select(Customer)
    cq = select(func.count(Customer.id))
    if search:
        term = f"%{search}%"
        q = q.where(
            or_(
                Customer.name.ilike(term),
                Customer.phone.ilike(term),
                Customer.email.ilike(term),
            )
        )
        cq = cq.where(
            or_(
                Customer.name.ilike(term),
                Customer.phone.ilike(term),
                Customer.email.ilike(term),
            )
        )
    if customer_type:
        q = q.where(Customer.type == customer_type)
        cq = cq.where(Customer.type == customer_type)
    if branch_id:
        q = q.where(Customer.branch_id == branch_id)
        cq = cq.where(Customer.branch_id == branch_id)
    total = int((await db.execute(cq)).scalar() or 0)
    conds = []
    if search:
        term = f"%{search}%"
        conds.append(
            or_(
                Customer.name.ilike(term),
                Customer.phone.ilike(term),
                Customer.email.ilike(term),
            )
        )
    if customer_type:
        conds.append(Customer.type == customer_type)
    if branch_id:
        conds.append(Customer.branch_id == branch_id)
    outstanding_total = float(
        (
            await db.execute(
                select(func.coalesce(func.sum(Customer.outstanding), 0)).where(and_(*conds) if conds else True)
            )
        ).scalar()
        or 0
    )
    wb_filter = and_(Customer.outstanding > 0, *conds) if conds else (Customer.outstanding > 0)
    with_balance = int((await db.execute(select(func.count(Customer.id)).where(wb_filter))).scalar() or 0)
    sort_expr = resolve_sort(
        sort_by,
        sort_order,
        {
            "name": Customer.name,
            "phone": Customer.phone,
            "email": Customer.email,
            "customer_type": Customer.type,
            "credit_limit": Customer.credit_limit,
            "outstanding": Customer.outstanding,
            "total_purchases": Customer.total_purchases,
            "created_at": Customer.created_at,
        },
        default_key="name",
        default_order="asc",
    )
    result = await db.execute(q.order_by(sort_expr).offset(sk).limit(lim))
    customers = result.scalars().all()
    await _attach_kam_names(db, customers)
    items = [serialize_customer(c) for c in customers]
    return paged(
        items,
        total,
        sk,
        lim,
        summary={"outstandingTotal": outstanding_total, "withBalanceCount": with_balance},
    )

@router.get("/{customer_id}", dependencies=[Depends(require_perm(*CUSTOMER_PICKER_READ))])
async def get_customer(customer_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Customer).where(Customer.id == customer_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Customer not found")
    await _attach_kam_names(db, [c])
    return serialize_customer(c)


@router.get("/{customer_id}/credit-ledger", dependencies=[Depends(require_perm("customers.view"))])
async def customer_credit_ledger(
    customer_id: str,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    """Append-only store-credit ledger for a customer (Sales Phase 1)."""
    exists = (
        await db.execute(select(Customer.id).where(Customer.id == customer_id))
    ).scalar_one_or_none()
    if not exists:
        raise HTTPException(404, "Customer not found")

    total = int(
        (
            await db.execute(
                select(func.count(CustomerCreditEntry.id)).where(
                    CustomerCreditEntry.customer_id == customer_id
                )
            )
        ).scalar()
        or 0
    )
    sk = normalize_skip(skip)
    lim = normalize_limit(limit)
    rows = (
        await db.execute(
            select(CustomerCreditEntry)
            .where(CustomerCreditEntry.customer_id == customer_id)
            .order_by(CustomerCreditEntry.created_at.desc(), CustomerCreditEntry.id.desc())
            .offset(sk)
            .limit(lim)
        )
    ).scalars().all()

    def _entry_type_label(t: str) -> str:
        labels = {
            "sale_debit": "Credit sale",
            "payment_debit": "Credit payment",
            "overpayment": "Overpayment credit",
            "return_credit": "Return credit",
            "return_void_revoke": "Return void (revoke)",
            "void_restore": "Payment void (restore)",
            "void_revoke": "Payment void (revoke)",
        }
        return labels.get(t, t.replace("_", " ").title())

    items = [{
        "id": e.id,
        "entryType": e.entry_type,
        "entryLabel": _entry_type_label(e.entry_type),
        "delta": round(float(e.delta or 0), 2),
        "balanceBefore": round(float(e.balance_before or 0), 2),
        "balanceAfter": round(float(e.balance_after or 0), 2),
        "sourceType": e.source_type,
        "sourceRef": e.source_ref,
        "sourceNumber": e.source_number,
        "notes": e.notes,
        "date": e.date,
        "createdBy": e.created_by,
        "createdAt": e.created_at.isoformat() if e.created_at else None,
    } for e in rows]
    return paged(items, total, sk, lim)

async def _validate_branch_id(branch_id: str, db: AsyncSession) -> None:
    result = await db.execute(select(Branch.id).where(Branch.id == branch_id))
    if result.scalar_one_or_none() is None:
        raise HTTPException(400, f"Unknown branch: {branch_id}")

def _compose_address_from_parts(data: Union[CustomerCreate, CustomerUpdate]) -> str:
    parts = [
        getattr(data, 'street1', None),
        getattr(data, 'street2', None),
        getattr(data, 'street3', None),
        getattr(data, 'city', None),
        getattr(data, 'state_province', None),
        getattr(data, 'country', None),
        getattr(data, 'postal_code', None),
    ]
    return ", ".join([p.strip() for p in parts if isinstance(p, str) and p.strip()])


async def _attach_kam_names(db: AsyncSession, customers: list[Customer]) -> None:
    """Resolve key_account_manager user ids → display names on each customer."""
    kam_ids = {
        c.key_account_manager
        for c in customers
        if getattr(c, "key_account_manager", None)
    }
    if not kam_ids:
        return
    rows = (
        await db.execute(select(User.id, User.name).where(User.id.in_(kam_ids)))
    ).all()
    name_by_id = {row.id: row.name for row in rows}
    for c in customers:
        kam = getattr(c, "key_account_manager", None)
        # Fall back to the raw value for legacy free-text entries.
        c._key_account_manager_name = name_by_id.get(kam) or kam


@router.post("/import", dependencies=[Depends(require_perm("customers.create"))])
async def import_customers(
    file: UploadFile = File(...),
    branch_id: str = Form(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(current_user),
):
    """Bulk import customers from an Excel file."""
    await _validate_branch_id(branch_id, db)
    await enforce_branch_access(branch_id, user=user, db=db)

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, detail=f"Failed to read Excel file: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(400, detail="Spreadsheet must have a header row and at least one data row")

    headers = [str(h).strip().lower() if h is not None else None for h in rows[0]]
    map_keys = {
        "name": "name", "customer name": "name", "phone": "phone", "email": "email",
        "gst reg no": "gst_in", "gst number": "gst_in", "gstin": "gst_in",
        "street 1": "street1", "street1": "street1", "street 2": "street2", "street2": "street2",
        "street 3": "street3", "street3": "street3", "city": "city",
        "state/province": "state_province", "state province": "state_province",
        "country": "country", "postal code": "postal_code", "postal_code": "postal_code",
        "credit limit": "credit_limit", "customer type": "customer_type",
        "key account manager": "key_account_manager", "credit terms": "credit_terms",
    }

    created = 0
    errors = []

    def as_text(value):
        if value is None:
            return None
        return value.strip() if isinstance(value, str) else str(value).strip()

    for idx, row in enumerate(rows[1:], start=2):
        try:
            data = {}
            for col_idx, cell in enumerate(row):
                key = headers[col_idx] if col_idx < len(headers) else None
                mapped = map_keys.get(key) if key else None
                if mapped:
                    data[mapped] = cell

            name = as_text(data.get("name"))
            if not name:
                raise ValueError("Customer name is required")
            required = {key: as_text(data.get(key)) for key in ("street1", "city", "country")}
            missing = [key for key, value in required.items() if not value]
            if missing:
                raise ValueError(f"Required field(s) missing: {', '.join(missing)}")

            customer_type = _normalize_customer_type(as_text(data.get("customer_type")) or "retail")
            raw_limit = data.get("credit_limit")
            if customer_type == "retail":
                credit_limit = 0.0
                credit_terms = None
            else:
                credit_limit = float(raw_limit if raw_limit not in (None, "") else 10000)
                credit_terms = as_text(data.get("credit_terms")) or None
            customer_id = str(uuid.uuid4())
            customer = Customer(
                id=customer_id,
                name=name,
                phone=as_text(data.get("phone")) or None,
                email=as_text(data.get("email")) or None,
                gstin=as_text(data.get("gst_in")) or None,
                branch_id=branch_id,
                credit_limit=credit_limit,
                type=customer_type,
                key_account_manager=as_text(data.get("key_account_manager")) or None,
                credit_terms=credit_terms,
                street1=required["street1"],
                street2=as_text(data.get("street2")) or None,
                street3=as_text(data.get("street3")) or None,
                city=required["city"],
                state_province=as_text(data.get("state_province")) or None,
                country=required["country"],
                postal_code=as_text(data.get("postal_code")) or None,
            )
            customer.address = _compose_address_from_parts(customer)
            customer.customer_code = _build_customer_code(customer_id)
            db.add(customer)
            await db.flush()
            await db.commit()
            created += 1
        except Exception as e:
            await db.rollback()
            errors.append({"row": idx, "error": str(e)})

    await db.commit()
    return {"created": created, "errors": errors}


@router.get("/import/template", dependencies=[Depends(require_perm("customers.create"))])
async def download_customer_import_template():
    """Generate an Excel template for bulk customer imports."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append([
        "Customer Name", "Phone", "Email", "GST Reg No", "Street 1", "Street 2", "Street 3",
        "City", "State/Province", "Country", "Postal Code",
        "Credit Limit", "Customer Type", "Key Account Manager", "Credit Terms",
    ])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customer_import_template.xlsx"'},
    )


@router.post("/", status_code=201, dependencies=[Depends(require_perm("customers.create"))])
async def create_customer(data: CustomerCreate, db: AsyncSession = Depends(get_db), user: User = Depends(current_user)):
    await _validate_branch_id(data.branch_id, db)
    await enforce_branch_access(data.branch_id, user=user, db=db)
    customer_type = _normalize_customer_type(data.customer_type)
    # Retail has no account credit facility — force limit/terms off.
    credit_limit = 0.0 if customer_type == "retail" else float(data.credit_limit or 0)
    credit_terms = None if customer_type == "retail" else (
        str(data.credit_terms) if data.credit_terms is not None else None
    )
    address = _compose_address_from_parts(data)
    customer_id = str(uuid.uuid4())
    customer_code = _build_customer_code(customer_id)
    c = Customer(id=customer_id, name=data.name, phone=data.phone,
                 email=data.email, address=address, customer_code=customer_code,
                 gstin=data.gst_in, branch_id=data.branch_id,
                 credit_limit=credit_limit, type=customer_type,
                 key_account_manager=(data.key_account_manager or None),
                 credit_terms=credit_terms,
                 street1=data.street1, street2=data.street2, street3=data.street3,
                 city=data.city, state_province=data.state_province,
                 country=data.country, postal_code=data.postal_code)
    db.add(c)
    await db.commit()
    return {"id": c.id, "message": "Customer created"}

@router.put("/{customer_id}", dependencies=[Depends(require_perm("customers.edit"))])
async def update_customer(customer_id: str, data: CustomerUpdate, db: AsyncSession = Depends(get_db), user: User = Depends(current_user)):
    result = await db.execute(select(Customer).where(Customer.id == customer_id))
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Customer not found")
    items = data.model_dump(exclude_unset=True)
    if "branch_id" in items:
        await _validate_branch_id(items["branch_id"], db)
        await enforce_branch_access(items["branch_id"], user=user, db=db)
    if "customer_type" in items:
        items["customer_type"] = _normalize_customer_type(items["customer_type"])
    if "credit_terms" in items and items["credit_terms"] is not None:
        items["credit_terms"] = str(items["credit_terms"])

    if any(k in items for k in [
        "street1", "street2", "street3", "city", "state_province", "country", "postal_code"
    ]):
        for k, v in items.items():
            if k in ["street1", "street2", "street3", "city", "state_province", "country", "postal_code"]:
                setattr(c, k, v)
        c.address = _compose_address_from_parts(data)

    for k, v in items.items():
        if k in ["street1", "street2", "street3", "city", "state_province", "country", "postal_code"]:
            continue
        setattr(c, _FIELD_TO_COLUMN.get(k, k), v)

    # After applying fields: retail never keeps an account limit / terms.
    effective_type = (c.type or "retail").strip().lower()
    if effective_type == "retail":
        c.credit_limit = 0.0
        c.credit_terms = None

    if not c.customer_code:
        c.customer_code = _build_customer_code(c.id)
    await db.commit()
    return {"message": "Updated"}
